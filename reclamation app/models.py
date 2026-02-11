from database import get_db, is_postgres

def _add_column_if_missing(db, table, column, col_def):
    if is_postgres():
        row = db.execute(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_name = ? AND column_name = ?
            """,
            (table, column),
        ).fetchone()
        if not row:
            db.execute(f"ALTER TABLE {table} ADD COLUMN {col_def}")
        return
    cols = db.execute(f"PRAGMA table_info({table})").fetchall()
    existing = {c["name"] for c in cols}
    if column not in existing:
        db.execute(f"ALTER TABLE {table} ADD COLUMN {col_def}")

def _province_from_code(code_value):
    if code_value is None:
        return None
    code_str = str(code_value).strip()
    if not code_str:
        return None
    first = code_str[0]
    return {
        "1": "ANTANANARIVO",
        "2": "ANTSIRANANA",
        "3": "FIANARANTSOA",
        "4": "MAHAJANGA",
        "5": "TOAMASINA",
        "6": "TOLIARA",
    }.get(first)

def _seed_bureaux_from_xlsx(db):
    import os
    try:
        from openpyxl import load_workbook
    except Exception:
        return

    path = os.path.join(os.path.abspath(os.path.dirname(__file__)), "codidue.xlsx")
    if not os.path.exists(path):
        return

    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            code = row[0]
            name = row[1] if len(row) > 1 else None
            if not code or not name:
                continue
            if str(code).strip().upper() == "CODIQUE":
                continue
            code_str = str(code).strip()
            name_str = str(name).strip()
            if not code_str or not name_str:
                continue
            province = _province_from_code(code_str)

            existing = db.execute(
                "SELECT id FROM bureaux WHERE code_bureau = ?",
                (code_str,),
            ).fetchone()
            if existing:
                db.execute(
                    "UPDATE bureaux SET nom_bureau = ?, province = ? WHERE code_bureau = ?",
                    (name_str, province, code_str),
                )
            else:
                db.execute(
                    "INSERT INTO bureaux (code_bureau, nom_bureau, province) VALUES (?, ?, ?)",
                    (code_str, name_str, province),
                )

def init_db():
    db = get_db()
    if is_postgres():
        db.executescript("""
        CREATE TABLE IF NOT EXISTS bureaux (
            id SERIAL PRIMARY KEY,
            code_bureau TEXT UNIQUE,
            nom_bureau TEXT,
            province TEXT
        );

        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE,
            password TEXT,
            role TEXT,
            bureau_id INTEGER,
            prenom TEXT,
            nom TEXT,
            matricule TEXT,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS types_reclamation (
            id SERIAL PRIMARY KEY,
            code TEXT UNIQUE,
            libelle TEXT,
            actif INTEGER DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS reclamations (
            id SERIAL PRIMARY KEY,
            numero_dossier TEXT,
            bureau_id INTEGER,
            user_id INTEGER,
            type_id INTEGER,
            numero_compte TEXT,
            nom_client TEXT,
            ancienne_valeur TEXT,
            nouvelle_valeur TEXT,
            motif TEXT,
            statut TEXT DEFAULT 'EN_ATTENTE',
            observation TEXT,
            archived INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS pieces_jointes (
            id SERIAL PRIMARY KEY,
            reclamation_id INTEGER,
            filename TEXT,
            original_name TEXT,
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS historique_statut (
            id SERIAL PRIMARY KEY,
            reclamation_id INTEGER,
            ancien_statut TEXT,
            nouveau_statut TEXT,
            observation TEXT,
            user_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
    else:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS bureaux (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code_bureau TEXT UNIQUE,
            nom_bureau TEXT,
            province TEXT
        );

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password TEXT,
            role TEXT,
            bureau_id INTEGER,
            prenom TEXT,
            nom TEXT,
            matricule TEXT,
            active INTEGER DEFAULT 1,
            created_at DATETIME DEFAULT (datetime('now', 'localtime'))
        );

        CREATE TABLE IF NOT EXISTS types_reclamation (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE,
            libelle TEXT,
            actif INTEGER DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS reclamations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_dossier TEXT,
            bureau_id INTEGER,
            user_id INTEGER,
            type_id INTEGER,
            numero_compte TEXT,
            nom_client TEXT,
            ancienne_valeur TEXT,
            nouvelle_valeur TEXT,
            motif TEXT,
            statut TEXT DEFAULT 'EN_ATTENTE',
            observation TEXT,
            archived INTEGER DEFAULT 0,
            created_at DATETIME DEFAULT (datetime('now', 'localtime')),
            updated_at DATETIME
        );

        CREATE TABLE IF NOT EXISTS pieces_jointes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reclamation_id INTEGER,
            filename TEXT,
            original_name TEXT,
            uploaded_at DATETIME DEFAULT (datetime('now', 'localtime'))
        );

        CREATE TABLE IF NOT EXISTS historique_statut (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reclamation_id INTEGER,
            ancien_statut TEXT,
            nouveau_statut TEXT,
            observation TEXT,
            user_id INTEGER,
            created_at DATETIME DEFAULT (datetime('now', 'localtime'))
        );
        """)

    if is_postgres():
        _add_column_if_missing(db, "users", "active", "active INTEGER DEFAULT 1")
        _add_column_if_missing(db, "users", "created_at", "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
    else:
        _add_column_if_missing(db, "users", "active", "active INTEGER DEFAULT 1")
        _add_column_if_missing(db, "users", "created_at", "created_at DATETIME DEFAULT (datetime('now', 'localtime'))")
    _add_column_if_missing(db, "users", "prenom", "prenom TEXT")
    _add_column_if_missing(db, "users", "nom", "nom TEXT")
    _add_column_if_missing(db, "users", "matricule", "matricule TEXT")

    _add_column_if_missing(db, "bureaux", "province", "province TEXT")

    _add_column_if_missing(db, "types_reclamation", "actif", "actif INTEGER DEFAULT 1")

    if is_postgres():
        _add_column_if_missing(db, "reclamations", "observation", "observation TEXT")
        _add_column_if_missing(db, "reclamations", "updated_at", "updated_at TIMESTAMP")
        _add_column_if_missing(db, "reclamations", "archived", "archived INTEGER DEFAULT 0")
        _add_column_if_missing(db, "reclamations", "reminder_requested_at", "reminder_requested_at TIMESTAMP")
        _add_column_if_missing(db, "reclamations", "reminder_disabled_until", "reminder_disabled_until TIMESTAMP")
        _add_column_if_missing(db, "reclamations", "reminder_auto_at", "reminder_auto_at TIMESTAMP")
        _add_column_if_missing(db, "reclamations", "reminder_last_sent_at", "reminder_last_sent_at TIMESTAMP")
        _add_column_if_missing(db, "reclamations", "reminder_auto_sent_at", "reminder_auto_sent_at TIMESTAMP")
    else:
        _add_column_if_missing(db, "reclamations", "observation", "observation TEXT")
        _add_column_if_missing(db, "reclamations", "updated_at", "updated_at DATETIME")
        _add_column_if_missing(db, "reclamations", "archived", "archived INTEGER DEFAULT 0")
        _add_column_if_missing(db, "reclamations", "reminder_requested_at", "reminder_requested_at DATETIME")
        _add_column_if_missing(db, "reclamations", "reminder_disabled_until", "reminder_disabled_until DATETIME")
        _add_column_if_missing(db, "reclamations", "reminder_auto_at", "reminder_auto_at DATETIME")
        _add_column_if_missing(db, "reclamations", "reminder_last_sent_at", "reminder_last_sent_at DATETIME")
        _add_column_if_missing(db, "reclamations", "reminder_auto_sent_at", "reminder_auto_sent_at DATETIME")

    # seed/ensure default types (insert missing, update labels if needed)
    default_types = [
        ("CHG_NOM", "Changement de nom"),
        ("CHG_TEL", "Changement de numero de telephone"),
        ("CHG_CIN", "Correction ou changement de numero CIN"),
        ("REG_MNT", "Regularisation de montant"),
        ("CHG_TYPE", "Changement de type de compte"),
        ("CHG_ADR", "Changement d'adresse"),
        ("CHG_EMAIL", "Changement d'email"),
        ("AUTRE", "Autre"),
    ]
    for code, libelle in default_types:
        if is_postgres():
            db.execute(
                """
                INSERT INTO types_reclamation (code, libelle, actif)
                VALUES (?, ?, 1)
                ON CONFLICT (code) DO UPDATE SET libelle = EXCLUDED.libelle, actif = 1
                """,
                (code, libelle),
            )
        else:
            db.execute(
                "INSERT OR IGNORE INTO types_reclamation (code, libelle, actif) VALUES (?, ?, 1)",
                (code, libelle),
            )
            db.execute(
                "UPDATE types_reclamation SET libelle = ?, actif = 1 WHERE code = ?",
                (libelle, code),
            )

    db.execute("UPDATE reclamations SET statut = 'TRAITEE' WHERE statut = 'VALIDEE'")
    db.execute("UPDATE historique_statut SET nouveau_statut = 'TRAITEE' WHERE nouveau_statut = 'VALIDEE'")
    db.execute("UPDATE historique_statut SET ancien_statut = 'TRAITEE' WHERE ancien_statut = 'VALIDEE'")

    _seed_bureaux_from_xlsx(db)

    db.commit()
    db.close()
